from disease import Disease
from tableReader import TableReader
import openpyxl

file = open("C:/Users/SAMSUNG/Desktop/jilbyong.txt", "r", encoding="UTF8")

def tableRead():
    tableReader = TableReader
    disease = []
    count = 0

    indexSetting = True #챕터 번호, 이름 받는 세션
    chunk = ""

    for line in file:
        if (indexSetting):
            indexSetting = tableReader.read(tableReader, line, disease)
        else:
            if (line[:8] == " 추이 및 현황"):
                diseaseInfo = ""
                risk = ""
                symptom = ""
                treat = ""
                prevent = ""

                #전처리 여기서 다 해야함
                if (chunk.find("질병정보") == -1 or chunk.find("위험요인") == -1):
                    diseaseInfo = ""
                else:
                    diseaseInfo = chunk[chunk.find("질병정보") + 5 : chunk.find("위험요인") - 1]
                    diseaseInfo = diseaseInfo.replace("\n", "")
                    diseaseInfo = diseaseInfo.replace("• ", ", ")
                if (chunk.find("위험요인") == -1 or chunk.find("증 상") == -1):
                    risk = ""
                else:
                    risk = chunk[chunk.find("위험요인") + 5 : chunk.find("증 상")]
                    risk = risk.replace("\n", "")
                    risk = risk.replace("• ", ", ")
                    if (risk[:2] == ", "):
                        risk = risk[2:]
                if (chunk.find("증 상") == -1 or chunk.find("치 료") == -1):
                    symptom = ""
                else:
                    symptom = chunk[chunk.find("증 상") + 6 : chunk.find("치 료")]
                    symptom = symptom.replace("\n", "")
                    symptom = symptom.replace("• ", ", ")
                    if (symptom[:2] == ", "):
                        symptom = symptom[2:]
                if (chunk.find("치 료") == -1 or chunk.find("예 방") == -1):
                    risk = ""
                else:
                    treat = chunk[chunk.find("치 료") + 6 : chunk.find("예 방")]
                    treat = treat.replace("\n", "")
                    treat = treat.replace("• ", ", ")
                    if (treat[:2] == ", "):
                        treat = treat[2:]
                if (chunk.find("예 방") == -1 or chunk.find(str(count)) == -1):
                    risk = ""
                else:
                    prevent = chunk[chunk.find("예 방") + 6 : chunk.find(str(count)) - 1]
                    prevent = prevent.replace("\n", "")
                    prevent = prevent.replace("• ", ", ")
                    if (prevent[:2] == ", "):
                        prevent = prevent[2:]
                disease[count - 1].addExplain(diseaseInfo, risk, symptom, treat, prevent)


                if count == 0:
                    count += 1
                    continue
                count += 1



                #초기화 - 다음 전처리를 준비하는것
                chunk = ""
            else:
                chunk += (line)
    


    print("show diseases")
    '''
    for d in disease:
        d.showAll()
    print("----------------")
    print(diseaseInfo)
    '''
    disease[50].showAll()
    wb = openpyxl.Workbook()
    wb.active.title = "diseaseList"
    sheet = wb["diseaseList"]


    sheet.cell(1, 1).value = "질병 이름"
    sheet.cell(1, 2).value = "질병 유형"
    sheet.cell(1, 3).value = "질병 분류"
    sheet.cell(1, 4).value = "질병 정보"
    sheet.cell(1, 5).value = "질병 원인"
    sheet.cell(1, 6).value = "질병 증상"
    sheet.cell(1, 7).value = "질병 치료 "
    sheet.cell(1, 8).value = "질병 예방"
    for i in range(100):
        sheet.cell(i + 2, 1).value = disease[i].diseaseName
        sheet.cell(i + 2, 2).value = disease[i].diseaseTag
        sheet.cell(i + 2, 3).value = disease[i].diseaseCategory
        sheet.cell(i + 2, 4).value = disease[i].diseaseInfo
        sheet.cell(i + 2, 5).value = disease[i].risk
        sheet.cell(i + 2, 6).value = disease[i].symptom
        sheet.cell(i + 2, 7).value = disease[i].treat
        sheet.cell(i + 2, 8).value = disease[i].prevent


        

    newFileName = "C:/Users/SAMSUNG/Desktop/결과물 폴더/질병 정보 정리.xlsx"
    wb.save(newFileName)



tableRead()
